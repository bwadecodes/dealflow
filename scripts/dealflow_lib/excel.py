"""Excel author — auditable workbook builder with firm-style formatting.

Standards (per v2 design spec):
- Source data tabs preserved unmodified
- Calculations reference source via formulas, never hardcoded values
- Method/Notes tab explains each calculation
- Input/formula/output cells visually distinguished by fill color
- Number formats and headers driven by firm-style profile
"""

from __future__ import annotations

from pathlib import Path
from typing import Any, Iterable

try:
    import openpyxl
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
except ImportError as _exc:
    raise ImportError(
        "openpyxl is required for Excel authoring. Install with: pip install openpyxl"
    ) from _exc

from . import firmstyle


def _hex_to_argb(hex_color: str) -> str:
    """openpyxl expects ARGB hex (8 chars). Accept '#rrggbb' or 'rrggbb'."""
    c = hex_color.lstrip("#")
    if len(c) == 6:
        return "FF" + c.upper()
    if len(c) == 8:
        return c.upper()
    raise ValueError(f"Invalid hex color: {hex_color}")


class ExcelAuthor:
    """Build firm-style Excel workbooks.

    Usage:
        author = ExcelAuthor(firm_style=load_profile())
        wb = author.new_workbook(title="Cohort Analysis — Acme Corp")
        author.add_source_tab(wb, "Source", rows=[["Cust", "MRR"], ["A", 100]])
        author.add_calc_tab(wb, "Calcs", header=["Metric", "Value"], rows=[...])
        author.add_method_tab(wb, "How this was built...")
        author.save(wb, Path("analysis.xlsx"))
    """

    def __init__(self, firm_style: dict[str, Any] | None = None) -> None:
        self.style = firm_style or firmstyle.load_profile()
        self.fmt = firmstyle.get_excel_formatting(self.style)

        self._header_fill = PatternFill(
            "solid", fgColor=_hex_to_argb(self.fmt["header_fill"])
        )
        self._header_font = Font(
            color=_hex_to_argb(self.fmt["header_font_color"]),
            bold=True,
        )
        self._input_fill = PatternFill(
            "solid", fgColor=_hex_to_argb(self.fmt["input_cell_color"])
        )
        self._output_fill = PatternFill(
            "solid", fgColor=_hex_to_argb(self.fmt["output_cell_color"])
        )
        side = Side(style=self.fmt.get("border_style", "thin"))
        self._cell_border = Border(left=side, right=side, top=side, bottom=side)

    # ------------- workbook lifecycle -------------

    def new_workbook(self, title: str | None = None) -> openpyxl.Workbook:
        wb = openpyxl.Workbook()
        # remove the default sheet — caller adds named tabs
        default = wb.active
        wb.remove(default)
        if title:
            wb.properties.title = title
        return wb

    def save(self, wb: openpyxl.Workbook, path: Path) -> Path:
        path = Path(path)
        path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(path)
        return path

    # ------------- tab builders -------------

    def add_source_tab(
        self,
        wb: openpyxl.Workbook,
        name: str,
        rows: Iterable[Iterable[Any]],
        *,
        first_row_is_header: bool = True,
    ) -> Any:
        """Source data tab — written exactly as given, no calculations."""
        ws = wb.create_sheet(name)
        for r_idx, row in enumerate(rows, start=1):
            for c_idx, value in enumerate(row, start=1):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)
                if first_row_is_header and r_idx == 1:
                    cell.fill = self._header_fill
                    cell.font = self._header_font
                    cell.alignment = Alignment(horizontal="left", vertical="center")
        if first_row_is_header:
            ws.freeze_panes = "A2"
        self._autosize(ws)
        return ws

    def add_calc_tab(
        self,
        wb: openpyxl.Workbook,
        name: str,
        header: list[str],
        rows: Iterable[Iterable[Any]],
        *,
        output_columns: list[int] | None = None,
        input_columns: list[int] | None = None,
    ) -> Any:
        """Calculation tab. Values starting with '=' are written as formulas.
        Cells in input_columns get input fill; output_columns get output fill."""
        ws = wb.create_sheet(name)
        # header
        for c_idx, h in enumerate(header, start=1):
            cell = ws.cell(row=1, column=c_idx, value=h)
            cell.fill = self._header_fill
            cell.font = self._header_font
            cell.alignment = Alignment(horizontal="left", vertical="center")
        # body
        for r_idx, row in enumerate(rows, start=2):
            for c_idx, value in enumerate(row, start=1):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)
                if input_columns and c_idx in input_columns:
                    cell.fill = self._input_fill
                if output_columns and c_idx in output_columns:
                    cell.fill = self._output_fill
        ws.freeze_panes = "A2"
        self._autosize(ws)
        return ws

    def add_summary_tab(
        self,
        wb: openpyxl.Workbook,
        name: str,
        bullets: list[str],
        *,
        title: str | None = None,
    ) -> Any:
        ws = wb.create_sheet(name, 0)  # insert at the front
        row = 1
        if title:
            ws.cell(row=row, column=1, value=title).font = Font(bold=True, size=14)
            row += 2
        for b in bullets:
            ws.cell(row=row, column=1, value=f"• {b}")
            row += 1
        ws.column_dimensions["A"].width = 100
        return ws

    def add_method_tab(
        self, wb: openpyxl.Workbook, methodology: str, *, name: str = "Method"
    ) -> Any:
        """Plain-text tab describing how the analysis was built. Long string
        wraps across rows for readability."""
        ws = wb.create_sheet(name)
        ws.cell(row=1, column=1, value="Methodology & Notes").font = Font(
            bold=True, size=14
        )
        for i, line in enumerate(methodology.splitlines() or [methodology], start=3):
            ws.cell(row=i, column=1, value=line)
        ws.column_dimensions["A"].width = 110
        return ws

    # ------------- formatting helpers -------------

    def apply_currency(self, cell) -> None:
        cell.number_format = self.fmt["number_format_currency"]

    def apply_percent(self, cell) -> None:
        cell.number_format = self.fmt["number_format_percent"]

    def _autosize(self, ws, max_width: int = 50) -> None:
        """Cheap column auto-fit based on observed string lengths."""
        for col_cells in ws.columns:
            length = 0
            col_letter = None
            for cell in col_cells:
                if col_letter is None:
                    col_letter = get_column_letter(cell.column)
                val = "" if cell.value is None else str(cell.value)
                if len(val) > length:
                    length = len(val)
            if col_letter:
                ws.column_dimensions[col_letter].width = min(
                    max(12, length + 2), max_width
                )
